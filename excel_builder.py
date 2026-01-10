
import openpyxl
from openpyxl.styles import Font, Alignment, Protection, Border, Side
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.utils import quote_sheetname, get_column_letter
import os
import sys
import io
import re


class UniqueNameGenerator:
    """
    Generates unique ShortNames (15 chars) and LongNames (23 chars) with duplicate detection.
    Uses numeric suffixes to differentiate duplicates.
    """
    
    def __init__(self):
        self.seen_shortnames = set()           # Track all used ShortNames
        self.full_to_short = {}                # "Eggplant Margherita Pizza (Personal)" → "Eggplant Margh1"
        self.full_to_long = {}                 # "Eggplant Margherita Pizza (Personal)" → "Eggplant Margherita Pi"
        
        # Separate tracking by entity type for clarity
        self.item_names = {}                   # Item full_name → shortname
        self.modifier_group_names = {}         # ModGroup full_name → shortname  
        self.submenu_names = {}                # Submenu full_name → shortname
    
    def clean_text(self, text):
        """Clean text of special characters before processing."""
        if not text:
            return ""
        cleaned = re.sub(r'[^a-zA-Z0-9\s\.,\'\-\(\)\&/<> \u00C0-\u00FF]', '', str(text))
        return cleaned.strip()
    
    def generate_unique_shortname(self, full_name, entity_type="item"):
        """
        Generate a unique 15-char ShortName for a given full name.
        
        Args:
            full_name: The original full name from AI
            entity_type: "item", "modifier_group", or "submenu"
            
        Returns:
            tuple: (unique_shortname, longname)
        """
        if not full_name:
            return "", ""
        
        # Clean the name first
        cleaned = self.clean_text(full_name)
        if not cleaned:
            return "", ""
        
        # Check if we already processed this exact full name
        if full_name in self.full_to_short:
            return self.full_to_short[full_name], self.full_to_long[full_name]
        
        # Generate LongName (23 chars)
        long_name = cleaned[:23]
        
        # Generate base ShortName (15 chars)
        base_short = cleaned[:15]
        
        # Check if this ShortName is unique
        if base_short not in self.seen_shortnames:
            short_name = base_short
        else:
            # Need to add numeric suffix for uniqueness
            short_name = self._get_unique_with_suffix(cleaned)
        
        # Store the mappings
        self.seen_shortnames.add(short_name)
        self.full_to_short[full_name] = short_name
        self.full_to_long[full_name] = long_name
        
        # Store in entity-specific dict
        if entity_type == "item":
            self.item_names[full_name] = short_name
        elif entity_type == "modifier_group":
            self.modifier_group_names[full_name] = short_name
        elif entity_type == "submenu":
            self.submenu_names[full_name] = short_name
        
        return short_name, long_name
    
    def _get_unique_with_suffix(self, cleaned_name):
        """
        Generate unique ShortName with numeric suffix.
        Tries: base14 + 1-9, then base13 + 10-99
        """
        # Try single digit suffix (1-9)
        base14 = cleaned_name[:14]
        for i in range(1, 10):
            candidate = f"{base14}{i}"
            if candidate not in self.seen_shortnames:
                return candidate
        
        # Try double digit suffix (10-99)
        base13 = cleaned_name[:13]
        for i in range(10, 100):
            candidate = f"{base13}{i}"
            if candidate not in self.seen_shortnames:
                return candidate
        
        # Fallback: triple digit (very unlikely)
        base12 = cleaned_name[:12]
        for i in range(100, 1000):
            candidate = f"{base12}{i}"
            if candidate not in self.seen_shortnames:
                return candidate
        
        # Last resort: just return with timestamp-like suffix
        import time
        return f"{cleaned_name[:10]}{int(time.time()) % 100000}"
    
    def lookup_shortname(self, full_name, entity_type=None):
        """
        Look up the ShortName for a given full name.
        Useful for cross-referencing (e.g., SubmenuItem referencing Item).
        """
        if full_name in self.full_to_short:
            return self.full_to_short[full_name]
        
        # Try truncated match (AI might have given slightly different name)
        cleaned = self.clean_text(full_name) if full_name else ""
        for stored_full, short in self.full_to_short.items():
            if self.clean_text(stored_full)[:15] == cleaned[:15]:
                return short
        
        # Not found - return truncated version (will likely fail validation but logged)
        return cleaned[:15] if cleaned else ""


class ExcelBuilder:
    def __init__(self):
        self.data = {
            "Item": [],
            "Submenu": [],
            "SubmenuItem": [],
            "ModifierGroup_Items": [],
            "ModifierGroup": [],
            "Category": [],
            "TaxGroup": [],
            "Menu": [],
            "MenuSubmenu": []
        }
        self.name_generator = UniqueNameGenerator()
        
        # Track item indices for Phase 3 modifier assignment
        self._item_indices = {}  # full_name → index in self.data["Item"]
        self._modifier_group_shortnames = set()  # Valid modifier group ShortNames

    def clean_text(self, text):
        if not text: return None
        cleaned = re.sub(r'[^a-zA-Z0-9\s\.,\'\-\(\)\&/<> \u00C0-\u00FF]', '', str(text))
        return cleaned.strip()

    def add_data(self, json_data: dict):
        """
        Main entry point for processing AI JSON data.
        Implements phased processing to handle dependencies correctly.
        """
        # Phase 0: Pre-process all names to generate unique ShortNames
        self._preprocess_names(json_data)
        
        # Phase 1: Create all items (regular + modifier items) with I-R empty
        self._create_items(json_data)
        
        # Phase 2: Create modifier groups (references Item ShortNames)
        self._create_modifier_groups(json_data)
        
        # Phase 3: Update items with modifier group assignments (I-R columns)
        self._assign_modifier_groups_to_items(json_data)
        
        # Phase 4: Create submenus and submenu items
        self._create_submenus(json_data)

    def _preprocess_names(self, json_data: dict):
        """
        Phase 0: Scan all names and generate unique ShortNames upfront.
        This ensures no duplicates across the entire dataset.
        """
        # Process modifier group names
        for mg in json_data.get("modifier_groups", []):
            raw_name = mg.get("name") or ""
            if raw_name:
                short, _ = self.name_generator.generate_unique_shortname(raw_name, "modifier_group")
                self._modifier_group_shortnames.add(short)
            
            # Process modifier items within each group
            for m_item in mg.get("items", []):
                m_raw = m_item.get("name") or ""
                if m_raw:
                    self.name_generator.generate_unique_shortname(m_raw, "item")
        
        # Process regular items
        for item in json_data.get("items", []):
            raw_name = item.get("name") or ""
            if raw_name:
                self.name_generator.generate_unique_shortname(raw_name, "item")
        
        # Process submenus
        for sm in json_data.get("submenus", []):
            raw_name = sm.get("name") or ""
            if raw_name:
                self.name_generator.generate_unique_shortname(raw_name, "submenu")

    def _create_items(self, json_data: dict):
        """
        Phase 1: Create all items (regular + modifier items).
        Modifier group assignments (columns I-R) are left empty at this stage.
        """
        mod_item_number_start = 20000
        mod_item_count = 0
        
        # First, add modifier items (from modifier_groups[].items[])
        for mg in json_data.get("modifier_groups", []):
            for m_item in mg.get("items", []):
                m_raw = m_item.get("name") or ""
                if not m_raw:
                    continue
                
                # Use pre-generated ShortName
                m_short = self.name_generator.lookup_shortname(m_raw, "item")
                m_long = self.name_generator.full_to_long.get(m_raw, m_raw[:23])
                
                try:
                    m_price = float(m_item.get("price", 0.0))
                except:
                    m_price = 0.0
                
                m_number = mod_item_number_start + mod_item_count
                mod_item_count += 1
                
                # Track index for later reference
                self._item_indices[m_raw] = len(self.data["Item"])
                
                # Add item row with empty modifier assignments (I-R = None)
                self.data["Item"].append([
                    m_number, m_short, m_long, "Standard", m_price, "Item Price",
                    None, None  # TaxGroup, Category
                ] + [None] * 10)  # ModifierGroup1-10 empty
        
        # Then, add regular items
        for item in json_data.get("items", []):
            raw_name = item.get("name") or ""
            if not raw_name:
                continue
            
            # Use pre-generated ShortName
            short_name = self.name_generator.lookup_shortname(raw_name, "item")
            long_name = self.name_generator.full_to_long.get(raw_name, raw_name[:23])
            
            try:
                price = item.get("price", 0.0)
                if isinstance(price, str):
                    price = float(price.replace('$', '').replace(',', ''))
            except:
                price = 0.0
            
            # Track index for Phase 3 modifier assignment
            self._item_indices[raw_name] = len(self.data["Item"])
            
            # Add item row with empty modifier assignments (filled in Phase 3)
            self.data["Item"].append([
                item.get("number"),
                short_name,
                long_name,
                "Standard",
                price,
                "Item Price",
                None,  # TaxGroupName
                None,  # CategoryName
            ] + [None] * 10)  # ModifierGroup1-10 empty for now

    def _create_modifier_groups(self, json_data: dict):
        """
        Phase 2: Create ModifierGroup_Items rows.
        References Item ShortNames for the modifier items.
        """
        for mg_idx, mg in enumerate(json_data.get("modifier_groups", [])):
            raw_name = mg.get("name") or ""
            if not raw_name:
                continue
            
            # Use pre-generated ShortName for the group
            mg_short = self.name_generator.lookup_shortname(raw_name, "modifier_group")
            mg_long = self.name_generator.full_to_long.get(raw_name, raw_name[:23])
            
            mg_num = mg.get("number")
            if not isinstance(mg_num, int):
                try:
                    mg_num = int(mg_num)
                except:
                    mg_num = 10000 + mg_idx
            
            if mg_num < 10000 or mg_num > 19999:
                mg_num = 10000 + (mg_idx * 10)
            
            items = mg.get("items", [])
            rows_needed = max(len(items), 6)
            
            for i in range(rows_needed):
                row_pos = i // 3
                col_pos = i % 3
                
                m_item_short = None
                m_price = None
                m_price_method = None
                
                if i < len(items):
                    m_item = items[i]
                    m_raw = m_item.get("name") or ""
                    
                    # Look up the pre-generated ShortName for this modifier item
                    m_item_short = self.name_generator.lookup_shortname(m_raw, "item")
                    m_price = "FORMULA_PRICE"
                    m_price_method = "Item Price"
                
                if i == 0:
                    # FIRST ROW - includes group header info
                    self.data["ModifierGroup_Items"].append([
                        mg_num, mg_short, mg_long, None, None, None, None,
                        mg_short,  # Column H: ModifierGroupName
                        m_item_short,  # Column I: ItemName (ShortName)
                        m_price, row_pos, col_pos, m_price_method
                    ])
                else:
                    # SPACER ROWS - only modifier item assignment
                    self.data["ModifierGroup_Items"].append([
                        None, None, None, None, None, None, None,
                        mg_short,  # Column H: ModifierGroupName
                        m_item_short,  # Column I: ItemName (ShortName)
                        m_price, row_pos, col_pos, m_price_method
                    ])

    def _assign_modifier_groups_to_items(self, json_data: dict):
        """
        Phase 3: Go back to Item data and fill columns I-R with modifier group ShortNames.
        Only assigns modifiers that exist in our modifier group data.
        """
        for item in json_data.get("items", []):
            raw_name = item.get("name") or ""
            if raw_name not in self._item_indices:
                continue
            
            item_idx = self._item_indices[raw_name]
            modifiers = item.get("modifiers", [])
            
            # Filter and convert modifiers to their ShortNames
            filtered_modifiers = []
            for mod in modifiers:
                if not mod:
                    continue
                
                # Look up the ShortName for this modifier group
                mod_short = self.name_generator.lookup_shortname(mod, "modifier_group")
                
                # Only include if it exists in our modifier group data
                if mod_short in self._modifier_group_shortnames:
                    filtered_modifiers.append(mod_short)
            
            # Update columns I-R (indices 8-17) in the item row
            for i, mod_short in enumerate(filtered_modifiers[:10]):
                self.data["Item"][item_idx][8 + i] = mod_short

    def _create_submenus(self, json_data: dict):
        """
        Phase 4: Create Submenu and SubmenuItem entries.
        SubmenuItem references use pre-generated ShortNames.
        """
        for sm in json_data.get("submenus", []):
            raw_name = sm.get("name") or ""
            if not raw_name:
                continue
            
            # Use pre-generated ShortName for submenu
            sm_short = self.name_generator.lookup_shortname(raw_name, "submenu")
            sm_long = self.name_generator.full_to_long.get(raw_name, raw_name[:23])
            
            self.data["Submenu"].append([
                sm.get("number"),
                sm_short,
                sm_long
            ])
            
            # Create SubmenuItem entries
            for idx, item_name in enumerate(sm.get("items", [])):
                row_pos = idx // 3
                col_pos = idx % 3
                
                # Look up the Item's ShortName (not the raw name from AI)
                item_short = self.name_generator.lookup_shortname(item_name, "item")
                
                self.data["SubmenuItem"].append([
                    sm_short,  # Column A: SubmenuName (ShortName)
                    "Item Button",
                    item_short,  # Column C: ItemName (ShortName)
                    "Item Price",
                    row_pos,
                    col_pos,
                    "FORMULA_PRICE"
                ])

    def update_instructions_tab(self, wb):
        if "Instructions" not in wb.sheetnames: return
        ws = wb["Instructions"]

        # 1. DELETE Rows FIRST to stabilize indices
        ws.delete_rows(7, 5)

        # 2. Add items 3-5 to IMPORTANT INSTRUCTIONS section
        ws.insert_rows(7, 3)
        ws.cell(row=7, column=1, value="3. All dropdowns are protected - select from the list only.")
        ws.cell(row=8, column=1, value="4. Fields left blank will be auto-generated based on database defaults.")
        ws.cell(row=9, column=1, value="5. Right-click row number -> Insert to add rows in ModifierGroup_Items.")

        # 2. FORMATTING
        for r in [21, 27, 31]:
             for cell in ws[r]:
                 if cell.value: cell.font = Font(bold=False)

        if ws["A26"].value: ws["A26"].font = Font(bold=True)

        ws["A31"].value = "5. Menu"
        ws["A31"].font = Font(bold=False)

        # 3. Clean Content ("TIPS")
        target_row = None
        target_col = None
        tips_content = ""

        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and "TIPS" in cell.value:
                    target_row = cell.row
                    target_col = cell.column
                    tips_content = cell.value
                    cell.value = cell.value.replace("TIPS", "").strip()
                    break
            if target_row: break

        if target_row:
             ws.cell(row=target_row-1, column=target_col, value="NOTE: To edit structure (Headers), go to Review > Unprotect Sheet.")

             parts = tips_content.split("TIPS")

             button_expl_lines = [
                parts[0].strip(),
                "",
                "BUTTON POSITION LOGIC:",
                "1. MenuSubmenu (ButtonPositionIndex):",
                "   - Sequential number (0, 1, 2...) determining sort order.",
                "",
                "2. SubmenuItem & ModifierGroup_Items (Row/Column):",
                "   - Layout: 3 Columns x 7 Rows per page.",
                "   - Rows: Start at 0 and ascend.",
                "   - Logic: Row = Index // 3, Column = Index % 3.",
                "",
                "NOTE: Ensure no duplicate position combinations within the same Group.",
                "",
                "MODIFIER GROUPS:",
                " - First Row: Enter Group Number, ShortName, LongName, AND First Item.",
                " - D,E,F,G (Min/Max/Free/FlowRequired): Only for group header rows.",
                " - Rows Below: Leave A-C BLANK. Enter Items in Cols I-M.",
                " - To Add More Items: Right Click Row Number -> Insert."
             ]

             if ws.max_row >= target_row:
                 pass

             for i, line in enumerate(button_expl_lines):
                 cell = ws.cell(row=target_row + i, column=target_col, value=line)
                 if "BUTTON POSITION LOGIC:" in line or "MODIFIER GROUPS:" in line:
                     cell.font = Font(bold=True)

    def get_template_path(self):
        filename = "Aloha_Import_Template_Generated.xlsx"
        if getattr(sys, 'frozen', False):
            if hasattr(sys, '_MEIPASS'):
                path = os.path.join(sys._MEIPASS, filename)
                if os.path.exists(path): return path
            path = os.path.join(os.path.dirname(sys.executable), filename)
            if os.path.exists(path): return path
        if os.path.exists(filename): return filename
        return filename

    def build_final(self, is_empty_template=False) -> bytes:
        template_path = self.get_template_path()
        wb = openpyxl.load_workbook(template_path)
        self.update_instructions_tab(wb)

        def get_list_formula(sheet, col_range):
            return f"{quote_sheetname(sheet)}!{col_range}"

        target_sheets = ["Item", "Submenu", "SubmenuItem", "ModifierGroup_Items", "Menu", "Category", "TaxGroup", "MenuSubmenu"]

        guide_notes = {
            "Item": ["Auto-ID", "Max 15 chars", "Max 23 chars (Full Name)", "Standard", "0.00", "Item Price", "Look up Tax Group", "Look up Category"],
            "SubmenuItem": ["Lookup Submenu", "Item Button", "Lookup Item", "Item Price", "0-6 (Row)", "0-2 (Col)", "=VLOOKUP_PRICE"],
            "ModifierGroup_Items": ["Auto-ID", "Max 15", "Max 23", "Min", "Max", "Free", "Flow", "Copy Name", "Lookup Item", "=VLOOKUP_PRICE", "0-6", "0-2", "Item Price"]
        }

        for sheet_name in target_sheets:
            if sheet_name not in wb.sheetnames: continue
            ws = wb[sheet_name]

            # 1. Clear Existing Data (Keep Header Row 1)
            if ws.max_row > 1:
                ws.delete_rows(2, amount=ws.max_row - 1)

            # 2. Add Guide Notes (Only to Empty Template)
            if is_empty_template:
                 notes = guide_notes.get(sheet_name)
                 if notes:
                     ws.append(notes)
                     for cell in ws[2]: cell.font = Font(italic=True)

            # 3. Handle Special Columns
            if sheet_name == "Category":
                owner_col = None
                for cell in ws[1]:
                     if cell.value == "OwnerName":
                         owner_col = cell.column
                         break
                if owner_col: ws.delete_cols(owner_col)

            if sheet_name == "SubmenuItem":
                header_cell = ws.cell(row=1, column=7, value="Price")
                ref_cell = ws.cell(row=1, column=6)
                header_cell.font = Font(name='Cambria', size=11, bold=True)
                if ref_cell.fill:
                    header_cell.fill = ref_cell.fill.copy()
                if ref_cell.alignment:
                    header_cell.alignment = ref_cell.alignment.copy()
                if ref_cell.border:
                    header_cell.border = ref_cell.border.copy()

            # 4. Insert Data (if not empty)
            if not is_empty_template:
                rows = self.data.get(sheet_name, [])
                if sheet_name in ["Category", "TaxGroup", "MenuSubmenu"]:
                    rows = []

                for idx, row_data in enumerate(rows):
                    curr_row = idx + 2

                    if sheet_name == "ModifierGroup_Items":
                        if len(row_data) > 9 and row_data[9] == "FORMULA_PRICE":
                             row_data[9] = f"=IFERROR(VLOOKUP(I{curr_row}, Item!$B:$E, 4, FALSE), 0.00)"

                    elif sheet_name == "SubmenuItem":
                         if len(row_data) > 6 and row_data[6] == "FORMULA_PRICE":
                             row_data[6] = f"=IFERROR(VLOOKUP(C{curr_row}, Item!$B:$E, 4, FALSE), 0.00)"

                    clean_row = []
                    for val in row_data:
                         if isinstance(val, str) and not val.startswith("="):
                             val = self.clean_text(val)
                         clean_row.append(val)

                    ws.append(clean_row)

                    if sheet_name == "ModifierGroup_Items" and clean_row[0] and "Right Click" in str(clean_row[0]):
                         ws.cell(row=curr_row, column=1).font = Font(italic=True, size=9)

            # 5. Protection (Robust Setup)
            ws.protection.sheet = False

            ws.protection = SheetProtection(
                sheet=True,
                formatCells=False,
                formatColumns=False,
                formatRows=False,
                insertColumns=False,
                insertRows=False,
                deleteColumns=False,
                deleteRows=False,
                sort=False,
                autoFilter=False,
                pivotTables=False,
                selectLockedCells=False,
                selectUnlockedCells=False
            )

            max_r = 5000
            for row in ws.iter_rows(min_row=2, max_row=max_r, max_col=30):
                for cell in row:
                    cell.protection = Protection(locked=False)

            # 6. Validations (STRICT)
            def add_strict_list(ws, formula, cell_range):
                dv = DataValidation(type="list", formula1=formula, allow_blank=True)
                dv.error = "Select from dropdown."
                dv.errorTitle = "Invalid Selection"
                dv.showErrorMessage = True
                ws.add_data_validation(dv)
                dv.add(cell_range)

            if sheet_name == "Item":
               add_strict_list(ws, get_list_formula('Category', '$B$2:$B$500'), f"H2:H{max_r}")
               add_strict_list(ws, get_list_formula('TaxGroup', '$B$2:$B$500'), f"G2:G{max_r}")
               add_strict_list(ws, get_list_formula('ModifierGroup_Items', '$B$2:$B$200'), f"I2:R{max_r}")

               dv_type = DataValidation(type="list", formula1='"Standard,Gift Card"', allow_blank=True)
               dv_type.showErrorMessage = True
               dv_type.error = "Select from dropdown."
               dv_type.errorTitle = "Invalid Selection"
               ws.add_data_validation(dv_type)
               dv_type.add(f"D2:D{max_r}")

               dv_pm = DataValidation(type="list", formula1='"Item Price,Price Level,Quantity Price,Ask For Price"', allow_blank=True)
               dv_pm.showErrorMessage = True
               dv_pm.error = "Select from dropdown."
               dv_pm.errorTitle = "Invalid Selection"
               ws.add_data_validation(dv_pm)
               dv_pm.add(f"F2:F{max_r}")

            elif sheet_name == "Category":
                dv_type = DataValidation(type="list", formula1='"General,Sales,Retail"', allow_blank=True)
                dv_type.showErrorMessage = True
                dv_type.error = "Select from dropdown."
                dv_type.errorTitle = "Invalid Selection"
                ws.add_data_validation(dv_type)
                dv_type.add(f"C2:C{max_r}")

            elif sheet_name == "ModifierGroup_Items":
                add_strict_list(ws, get_list_formula('Item', '$B$2:$B$2000'), f"I2:I{max_r}")

                dv_pm = DataValidation(type="list", formula1='"Item Price,Button Price"', allow_blank=True)
                dv_pm.showErrorMessage = True
                ws.add_data_validation(dv_pm)
                dv_pm.add(f"M2:M{max_r}")

                dv_price = DataValidation(type="custom", formula1='=M2<>"Item Price"')
                dv_price.error = "Price is linked to Item Default. Change Method to edit."
                dv_price.showErrorMessage = True
                ws.add_data_validation(dv_price)
                dv_price.add(f"J2:J{max_r}")

                dv_flow = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
                dv_flow.showErrorMessage = True
                dv_flow.error = "Select from dropdown."
                dv_flow.errorTitle = "Invalid Selection"
                ws.add_data_validation(dv_flow)
                dv_flow.add(f"G2:G{max_r}")

                dv_nums = DataValidation(
                    type="custom",
                    formula1='=OR(ISBLANK($A2), AND(NOT(ISBLANK($A2)), NOT(ISBLANK($B2)), NOT(ISBLANK($C2))))'
                )
                dv_nums.errorTitle = "Group Row Required"
                dv_nums.error = "These fields are only for rows with Number, ShortName, and LongName filled in."
                dv_nums.showErrorMessage = True
                ws.add_data_validation(dv_nums)
                dv_nums.add(f"D2:F{max_r}")

                dv_lock = DataValidation(type="custom", formula1='=ISBLANK($I2)')
                dv_lock.error = "Number, ShortName, LongName must be blank for item rows (where ItemName is filled)."
                dv_lock.showErrorMessage = True
                ws.add_data_validation(dv_lock)
                dv_lock.add(f"A2:C{max_r}")

            elif sheet_name == "SubmenuItem":
                add_strict_list(ws, get_list_formula('Submenu', '$B$2:$B$500'), f"A2:A{max_r}")
                add_strict_list(ws, get_list_formula('Item', '$B$2:$B$2000'), f"C2:C{max_r}")

                dv_type = DataValidation(type="list", formula1='"Item Button,PLU Button"', allow_blank=True)
                dv_type.showErrorMessage = True
                dv_type.error = "Select from dropdown."
                dv_type.errorTitle = "Invalid Selection"
                ws.add_data_validation(dv_type)
                dv_type.add(f"B2:B{max_r}")

                dv_pm = DataValidation(type="list", formula1='"Item Price,Button Price,Price Level"', allow_blank=True)
                dv_pm.showErrorMessage = True
                dv_pm.error = "Select from dropdown."
                dv_pm.errorTitle = "Invalid Selection"
                ws.add_data_validation(dv_pm)
                dv_pm.add(f"D2:D{max_r}")

                dv_p = DataValidation(type="custom", formula1='=D2<>"Item Price"')
                dv_p.error = "Price is linked to Item Default."
                dv_p.showErrorMessage = True
                ws.add_data_validation(dv_p)
                dv_p.add(f"G2:G{max_r}")

            elif sheet_name == "MenuSubmenu":
                add_strict_list(ws, get_list_formula('Menu', '$B$2:$B$100'), f"A2:A{max_r}")
                add_strict_list(ws, get_list_formula('Submenu', '$B$2:$B$500'), f"B2:B{max_r}")

        # Apply Comment Logic for ModifierGroup_Items
        if "ModifierGroup_Items" in wb.sheetnames:
            ws = wb["ModifierGroup_Items"]
            for row_idx in range(2, ws.max_row + 1):
                if row_idx > 2:
                    current_cell_A = ws.cell(row=row_idx, column=1)
                    prev_cell_A = ws.cell(row=row_idx - 1, column=1)
                    current_cell_H = ws.cell(row=row_idx, column=8)

                    if current_cell_A.value is None and prev_cell_A.value is not None and current_cell_H.value is not None:
                        current_cell_A.comment = Comment("Right Click Number to Insert Rows", "System")

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def build_excel(self) -> bytes:
        return self.build_final(is_empty_template=False)

    def build_empty_template(self) -> bytes:
        return self.build_final(is_empty_template=True)

if __name__ == "__main__":
    print("Testing Consolidated Builder...")
    builder = ExcelBuilder()

    dummy_data = {
        "categories": [{"number": 1, "name": "Food"}, {"number": 2, "name": "Drinks"}],
        "items": [
            {"number": 100, "name": "Burger", "price": 10.0, "category": "Food", "modifiers": ["Sides"]},
            {"number": 101, "name": "Fries", "price": 5.0, "category": "Food"}
        ],
        "modifier_groups": [
            {"number": 1000, "name": "Sides", "min": 1, "max": 1, "items": [{"name": "Fries", "price": 0, "number":101}]}
        ],
        "submenus": [{"number": 200, "name": "Entrees", "items": ["Burger"]}]
    }
    builder.add_data(dummy_data)

    data = builder.build_excel()
    with open("Aloha_Import_Template_Consolidated.xlsx", "wb") as f:
        f.write(data)
    print("Success! Created 'Aloha_Import_Template_Consolidated.xlsx'.")

    empty_data = builder.build_empty_template()
    with open("Aloha_Import_Template_Consolidated_Empty.xlsx", "wb") as f:
        f.write(empty_data)
    print("Success! Created 'Aloha_Import_Template_Consolidated_Empty.xlsx'.")
